import os
import re
import pandas as pd
from glob import glob
from html.parser import HTMLParser
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURACAO
# ============================================================

PASTA = r"C:\Users\Samuel\Desktop\CERTIF\CERTIFICADOS"
ARQUIVO_SAIDA = "CONSOLIDADO_FINAL.xlsx"

# Arquivos auxiliares de AVP devem ser salvos na mesma pasta com nomes como:
# JAN 2026.xlsx, FEV 2026.xls, MAR 2026 - atualizado.xlsx etc.
# Esses arquivos NAO entram como relatorio principal. Eles servem apenas para trazer
# o Nome do AVP por Protocolo e preencher a coluna AGR no consolidado/dashboard.
MESES_AVP = {
    "JAN", "FEV", "MAR", "ABR", "MAI", "JUN",
    "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"
}


# ============================================================
# LEITOR HTML SEM DEPENDER DE LXML/BS4
# ============================================================

class TabelaHTMLParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.tabelas = []
        self.tabela_atual = None
        self.linha_atual = None
        self.celula_atual = None
        self.dentro_td = False

    def handle_starttag(self, tag, attrs):
        tag = tag.lower()

        if tag == "table":
            self.tabela_atual = []

        elif tag == "tr" and self.tabela_atual is not None:
            self.linha_atual = []

        elif tag in ("td", "th") and self.linha_atual is not None:
            self.celula_atual = []
            self.dentro_td = True

        elif tag == "br" and self.dentro_td and self.celula_atual is not None:
            self.celula_atual.append("\n")

    def handle_endtag(self, tag):
        tag = tag.lower()

        if tag in ("td", "th") and self.dentro_td:
            texto = "".join(self.celula_atual).strip()
            texto = re.sub(r"[ \t]+", " ", texto)
            self.linha_atual.append(texto)
            self.celula_atual = None
            self.dentro_td = False

        elif tag == "tr" and self.linha_atual is not None:
            if any(str(x).strip() for x in self.linha_atual):
                self.tabela_atual.append(self.linha_atual)
            self.linha_atual = None

        elif tag == "table" and self.tabela_atual is not None:
            self.tabelas.append(self.tabela_atual)
            self.tabela_atual = None

    def handle_data(self, data):
        if self.dentro_td and self.celula_atual is not None:
            self.celula_atual.append(data)


def ler_html_manual(caminho):
    with open(caminho, "r", encoding="utf-8-sig", errors="ignore") as f:
        html = f.read()

    parser = TabelaHTMLParser()
    parser.feed(html)

    if not parser.tabelas:
        raise Exception("Nenhuma tabela HTML encontrada.")

    tabela = max(parser.tabelas, key=len)

    cabecalho = tabela[0]
    linhas = tabela[1:]

    max_cols = max(len(cabecalho), max(len(l) for l in linhas))
    cabecalho = cabecalho + [f"Coluna_{i + 1}" for i in range(len(cabecalho), max_cols)]

    linhas_ajustadas = []
    for linha in linhas:
        linha = linha + [""] * (max_cols - len(linha))
        linhas_ajustadas.append(linha[:max_cols])

    return pd.DataFrame(linhas_ajustadas, columns=cabecalho[:max_cols])


def ler_arquivo(caminho):
    ext = os.path.splitext(caminho)[1].lower()

    # 1) XLSX real
    if ext == ".xlsx":
        try:
            return pd.read_excel(caminho, engine="openpyxl")
        except Exception:
            pass

    # 2) XLS antigo real
    if ext == ".xls":
        try:
            return pd.read_excel(caminho, engine="xlrd")
        except Exception:
            pass

    # 3) HTML disfarcado de XLS via pandas
    try:
        tabelas = pd.read_html(caminho, encoding="utf-8")
        if tabelas:
            df = tabelas[0]

            primeira_linha = [str(x).strip().lower() for x in df.iloc[0].tolist()]
            if "data" in primeira_linha and "nome" in primeira_linha:
                df.columns = df.iloc[0]
                df = df.iloc[1:].reset_index(drop=True)

            return df
    except Exception:
        pass

    # 4) HTML disfarcado de XLS via parser interno
    try:
        return ler_html_manual(caminho)
    except Exception as e:
        raise Exception(f"Nao foi possivel identificar/ler o formato do arquivo. Detalhe: {e}")


# ============================================================
# EXTRACOES
# ============================================================

def extrair_nome_cpf_parceiro(valor):
    texto = "" if pd.isna(valor) else str(valor)
    texto = texto.replace("\r", "\n")
    texto = re.sub(r"\s+", " ", texto).strip()

    nome = texto
    cpf_cnpj = ""
    parceiro = ""

    m_cpf = re.search(r"CPF\s*/?\s*CNPJ\s*:\s*(.*?)(?:Parceiro\s*:|$)", texto, flags=re.I)
    if m_cpf:
        cpf_cnpj = m_cpf.group(1).strip()
        nome = texto[:m_cpf.start()].strip()

    m_parceiro = re.search(r"Parceiro\s*:\s*(.*)$", texto, flags=re.I)
    if m_parceiro:
        parceiro = m_parceiro.group(1).strip()

    return pd.Series([nome, cpf_cnpj, parceiro])


def extrair_pedido_emissor(valor):
    texto = "" if pd.isna(valor) else str(valor)
    texto = texto.replace("\r", "\n")
    texto = re.sub(r"\s+", " ", texto).strip()

    pedido = ""
    emissor = ""

    m = re.match(r"^(\d+)\s*(.*)$", texto)
    if m:
        pedido = m.group(1).strip()
        emissor = m.group(2).strip()
    else:
        partes = texto.split(" ", 1)
        pedido = partes[0].strip() if len(partes) > 0 else ""
        emissor = partes[1].strip() if len(partes) > 1 else ""

    return pd.Series([pedido, emissor])


def localizar_coluna(df, nome_procurado, posicao_fallback):
    for col in df.columns:
        if str(col).strip().lower() == nome_procurado.lower():
            return col

    if df.shape[1] > posicao_fallback:
        return df.columns[posicao_fallback]

    raise Exception(
        f"Nao encontrei a coluna {nome_procurado} e o arquivo nao possui a posicao esperada."
    )


def normalizar_texto_coluna(valor):
    texto = "" if pd.isna(valor) else str(valor)
    texto = texto.strip().lower()
    texto = re.sub(r"[áàãâä]", "a", texto)
    texto = re.sub(r"[éèêë]", "e", texto)
    texto = re.sub(r"[íìîï]", "i", texto)
    texto = re.sub(r"[óòõôö]", "o", texto)
    texto = re.sub(r"[úùûü]", "u", texto)
    texto = re.sub(r"ç", "c", texto)
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def localizar_coluna_flexivel(df, nomes_possiveis, obrigatoria=True):
    mapa = {normalizar_texto_coluna(col): col for col in df.columns}

    for nome in nomes_possiveis:
        chave = normalizar_texto_coluna(nome)
        if chave in mapa:
            return mapa[chave]

    # busca parcial para variações como "Nº Protocolo", "Nome AVP", etc.
    for chave, col in mapa.items():
        for nome in nomes_possiveis:
            nome_norm = normalizar_texto_coluna(nome)
            if nome_norm and nome_norm in chave:
                return col

    if obrigatoria:
        raise Exception(f"Nao encontrei nenhuma destas colunas: {', '.join(nomes_possiveis)}")
    return None


def normalizar_protocolo(valor):
    if pd.isna(valor):
        return ""
    texto = str(valor).strip()
    if texto.endswith(".0"):
        texto = texto[:-2]
    texto = re.sub(r"\s+", "", texto)
    return texto.upper()


def nome_arquivo_eh_base_avp(nome_arquivo):
    nome_sem_ext = os.path.splitext(os.path.basename(nome_arquivo))[0].upper().strip()
    partes = re.split(r"[\s_\-]+", nome_sem_ext)
    if len(partes) < 2:
        return False

    mes = partes[0]
    ano_encontrado = any(re.fullmatch(r"20\d{2}", parte) for parte in partes[1:])
    return mes in MESES_AVP and ano_encontrado


def carregar_base_avp(arquivos_avp):
    lista = []

    for arquivo in arquivos_avp:
        nome_arquivo = os.path.basename(arquivo)
        print(f"Lendo base AVP: {nome_arquivo}")

        try:
            df_avp = ler_arquivo(arquivo)
            df_avp = df_avp.dropna(how="all").reset_index(drop=True)

            if df_avp.empty:
                print(f"Base AVP vazia ignorada: {nome_arquivo}")
                continue

            col_protocolo = localizar_coluna_flexivel(
                df_avp,
                ["Protocolo", "N Protocolo", "Numero Protocolo", "Nº Protocolo", "Num Protocolo"],
            )
            col_avp = localizar_coluna_flexivel(
                df_avp,
                ["Nome do AVP", "Nome AVP", "AVP", "AGR", "Nome do AGR", "Nome AGR"],
            )

            temp = df_avp[[col_protocolo, col_avp]].copy()
            temp.columns = ["Protocolo_Normalizado", "Nome do AVP"]
            temp["Protocolo_Normalizado"] = temp["Protocolo_Normalizado"].apply(normalizar_protocolo)
            temp["Nome do AVP"] = temp["Nome do AVP"].fillna("").astype(str).str.strip()
            temp["Arquivo_AVP"] = nome_arquivo
            temp = temp[temp["Protocolo_Normalizado"].astype(str).str.len() > 0]
            temp = temp[temp["Nome do AVP"].astype(str).str.len() > 0]

            lista.append(temp)
            print(f"OK -> base AVP {nome_arquivo} | Protocolos: {len(temp)}")

        except Exception as e:
            print(f"ERRO -> base AVP {nome_arquivo}: {e}")

    if not lista:
        return pd.DataFrame(columns=["Protocolo_Normalizado", "Nome do AVP", "Arquivo_AVP"])

    base_avp = pd.concat(lista, ignore_index=True)
    base_avp = base_avp.drop_duplicates(subset=["Protocolo_Normalizado"], keep="last")
    return base_avp


def aplicar_avp_no_consolidado(df, base_avp):
    df = df.copy()

    col_protocolo = localizar_coluna_flexivel(
        df,
        ["Protocolo", "N Protocolo", "Numero Protocolo", "Nº Protocolo", "Num Protocolo"],
        obrigatoria=False,
    )

    if col_protocolo is None:
        print("AVISO: Nao encontrei coluna Protocolo no consolidado. AGR/AVP nao foi aplicado.")
        df["Protocolo_Normalizado"] = ""
        df["Nome do AVP"] = ""
        df["AGR"] = ""
        return df

    df["Protocolo_Normalizado"] = df[col_protocolo].apply(normalizar_protocolo)

    if base_avp.empty:
        print("AVISO: Nenhuma base AVP foi carregada. AGR ficara vazio quando nao houver AGR antigo.")
        if "Nome do AVP" not in df.columns:
            df["Nome do AVP"] = ""
        if "AGR" not in df.columns:
            df["AGR"] = df["Nome do AVP"]
        return df

    df = df.merge(base_avp, on="Protocolo_Normalizado", how="left")

    # Se ja existir uma coluna AGR antiga, ela fica preservada em AGR_Original.
    if "AGR" in df.columns:
        df["AGR_Original"] = df["AGR"]

    df["Nome do AVP"] = df["Nome do AVP"].fillna("").astype(str).str.strip()
    df["AGR"] = df["Nome do AVP"]

    protocolos_com_avp = (df["Nome do AVP"].astype(str).str.len() > 0).sum()
    print(f"AVP aplicado no consolidado. Linhas com AVP encontrado: {protocolos_com_avp}")

    return df


def coluna_excel(df, nome_coluna):
    if nome_coluna not in df.columns:
        return None
    return get_column_letter(list(df.columns).index(nome_coluna) + 1)


def normalizar_documento(valor):
    return re.sub(r"\D", "", "" if pd.isna(valor) else str(valor))


def identificar_tipo_documento(valor):
    doc = normalizar_documento(valor)
    if len(doc) == 11:
        return "CPF"
    if len(doc) == 14:
        return "CNPJ"
    return "NAO IDENTIFICADO"


def converter_valor(valor):
    if pd.isna(valor):
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor)
    texto = texto.replace("R$", "").replace(" ", "").strip()
    texto = texto.replace(".", "").replace(",", ".")
    texto = re.sub(r"[^0-9.\-]", "", texto)

    try:
        return float(texto)
    except Exception:
        return 0.0


def preparar_campos_indicadores(df):
    df = df.copy()
    df["Data_Tratada"] = pd.to_datetime(df.get("Data"), dayfirst=True, errors="coerce")
    df["Periodo"] = df["Data_Tratada"].dt.strftime("%m/%Y")
    df["Ano"] = df["Data_Tratada"].dt.year
    df["Documento_Normalizado"] = df.get("CPF_CNPJ", "").apply(normalizar_documento)
    df["Tipo_Documento"] = df.get("CPF_CNPJ", "").apply(identificar_tipo_documento)
    df["Valor_Numerico"] = df.get("Valor", "").apply(converter_valor)
    return df


def criar_lista_renovacao(df):
    base = df[
        (df["Ano"] == 2025)
        & (df["Documento_Normalizado"].astype(str).str.len() > 0)
    ].copy()
    renovados_2026 = set(
        df.loc[
            (df["Ano"] == 2026)
            & (df["Documento_Normalizado"].astype(str).str.len() > 0),
            "Documento_Normalizado",
        ].astype(str)
    )

    if base.empty:
        return pd.DataFrame(
            columns=[
                "Documento_Normalizado",
                "CPF_CNPJ",
                "Tipo_Documento",
                "Nome_Separado",
                "Ultima_Data_2025",
                "Qtd_Certificados_2025",
                "Valor_Total_2025",
                "Ultimo_Modelo_2025",
                "Ultimo_Vendedor_2025",
                "Status_Renovacao",
            ]
        )

    base = base.sort_values(["Documento_Normalizado", "Data_Tratada"])
    agrupado = (
        base.groupby("Documento_Normalizado", as_index=False)
        .agg(
            CPF_CNPJ=("CPF_CNPJ", "last"),
            Tipo_Documento=("Tipo_Documento", "last"),
            Nome_Separado=("Nome_Separado", "last"),
            Ultima_Data_2025=("Data_Tratada", "max"),
            Qtd_Certificados_2025=("Pedido_Numero", "count"),
            Valor_Total_2025=("Valor_Numerico", "sum"),
            Ultimo_Modelo_2025=("Modelo", "last"),
            Ultimo_Vendedor_2025=("Vendedor", "last"),
        )
    )
    agrupado["Status_Renovacao"] = agrupado["Documento_Normalizado"].astype(str).apply(
        lambda doc: "RENOVOU EM 2026" if doc in renovados_2026 else "NAO RENOVOU EM 2026"
    )
    return agrupado[agrupado["Status_Renovacao"] == "NAO RENOVOU EM 2026"].copy()


def ajustar_larguras(ws, max_width=42):
    for coluna in ws.columns:
        letra = coluna[0].column_letter
        maior = 0
        for celula in coluna:
            valor = "" if celula.value is None else str(celula.value)
            maior = max(maior, min(len(valor) + 2, max_width))
        ws.column_dimensions[letra].width = max(10, maior)


def aplicar_estilo_tabela(ws, nome_tabela):
    used = ws.calculate_dimension()
    if used == "A1:A1" and ws["A1"].value is None:
        return
    tabela = Table(displayName=nome_tabela, ref=used)
    estilo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)
    ws.freeze_panes = "A2"
    ajustar_larguras(ws)


def montar_aba_indicadores(writer, df):
    wb = writer.book
    ws = wb.create_sheet("INDICADORES COMERCIAIS")

    periodos = (
        df.dropna(subset=["Data_Tratada"])
        .sort_values("Data_Tratada")["Periodo"]
        .dropna()
        .drop_duplicates()
        .tolist()
    )
    if not periodos:
        periodos = ["SEM DATA"]

    ws_periodos = wb.create_sheet("AUX_PERIODOS")
    ws_periodos.sheet_state = "hidden"
    ws_periodos["A1"] = "Periodos"
    for idx, periodo in enumerate(periodos, start=2):
        ws_periodos.cell(idx, 1).value = periodo

    ws["A1"] = "INDICADORES COMERCIAIS"
    ws["A3"] = "Filtro lateral"
    ws["A4"] = "Periodo"
    ws["B4"] = periodos[-1]

    validacao = DataValidation(
        type="list",
        formula1=f"=AUX_PERIODOS!$A$2:$A${len(periodos) + 1}",
        allow_blank=False,
    )
    ws.add_data_validation(validacao)
    validacao.add(ws["B4"])

    indicadores = [
        ("Total de certificados", '=COUNTIF(CONSOLIDADO!$P:$P,$B$4)'),
        ("Valor total", '=SUMIF(CONSOLIDADO!$P:$P,$B$4,CONSOLIDADO!$T:$T)'),
        ("Ticket medio", '=IFERROR(B8/B7,0)'),
        ("Clientes unicos", '=COUNTA(UNIQUE(FILTER(CONSOLIDADO!$R:$R,(CONSOLIDADO!$P:$P=$B$4)*(CONSOLIDADO!$R:$R<>""))))'),
        ("Certificados CPF", '=COUNTIFS(CONSOLIDADO!$P:$P,$B$4,CONSOLIDADO!$S:$S,"CPF")'),
        ("Certificados CNPJ", '=COUNTIFS(CONSOLIDADO!$P:$P,$B$4,CONSOLIDADO!$S:$S,"CNPJ")'),
    ]

    ws["A6"] = "Indicador"
    ws["B6"] = "Resultado"
    for linha, (rotulo, formula) in enumerate(indicadores, start=7):
        ws.cell(linha, 1).value = rotulo
        ws.cell(linha, 2).value = formula

    ws["D6"] = "Origem"
    ws["E6"] = "Quantidade"
    origens = sorted([x for x in df["Origem"].dropna().astype(str).unique() if x.strip()])
    for linha, origem in enumerate(origens, start=7):
        ws.cell(linha, 4).value = origem
        ws.cell(linha, 5).value = f'=COUNTIFS(CONSOLIDADO!$P:$P,$B$4,CONSOLIDADO!$H:$H,D{linha})'

    ws["G6"] = "Vendedor"
    ws["H6"] = "Quantidade no periodo"
    vendedores = (
        df.groupby("Vendedor", dropna=True)["Pedido_Numero"]
        .count()
        .sort_values(ascending=False)
        .head(20)
        .index.astype(str)
        .tolist()
    )
    for linha, vendedor in enumerate(vendedores, start=7):
        ws.cell(linha, 7).value = vendedor
        ws.cell(linha, 8).value = f'=COUNTIFS(CONSOLIDADO!$P:$P,$B$4,CONSOLIDADO!$F:$F,G{linha})'

    ws["J6"] = "Modelo"
    ws["K6"] = "Quantidade no periodo"
    modelos = (
        df.groupby("Modelo", dropna=True)["Pedido_Numero"]
        .count()
        .sort_values(ascending=False)
        .head(20)
        .index.astype(str)
        .tolist()
    )
    for linha, modelo in enumerate(modelos, start=7):
        ws.cell(linha, 10).value = modelo
        ws.cell(linha, 11).value = f'=COUNTIFS(CONSOLIDADO!$P:$P,$B$4,CONSOLIDADO!$C:$C,J{linha})'

    # Ranking AGR/AVP: agora o AGR considerado no dashboard e o Nome do AVP
    ws["M6"] = "AGR / Nome do AVP"
    ws["N6"] = "Quantidade no periodo"
    col_periodo = coluna_excel(df, "Periodo")
    col_agr = coluna_excel(df, "AGR")
    avps = []
    if "AGR" in df.columns:
        avps = (
            df.loc[df["AGR"].fillna("").astype(str).str.strip() != ""]
            .groupby("AGR", dropna=True)["Pedido_Numero"]
            .count()
            .sort_values(ascending=False)
            .head(20)
            .index.astype(str)
            .tolist()
        )
    for linha, avp in enumerate(avps, start=7):
        ws.cell(linha, 13).value = avp
        if col_periodo and col_agr:
            ws.cell(linha, 14).value = f'=COUNTIFS(CONSOLIDADO!${col_periodo}:${col_periodo},$B$4,CONSOLIDADO!${col_agr}:${col_agr},M{linha})'
        else:
            ws.cell(linha, 14).value = ""

    ws["A15"] = "Ao alterar o periodo em B4, os indicadores desta aba sao recalculados no Excel."

    header_fill = PatternFill("solid", fgColor="1F4E78")
    title_fill = PatternFill("solid", fgColor="D9EAF7")
    white_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")

    ws["A1"].font = Font(size=16, bold=True, color="1F4E78")
    ws["A3"].font = Font(bold=True, color="1F4E78")
    ws["A3"].fill = title_fill
    ws["A4"].font = Font(bold=True)
    ws["B4"].fill = PatternFill("solid", fgColor="FFF2CC")

    for row in (6,):
        for col in range(1, 15):
            cell = ws.cell(row, col)
            if cell.value:
                cell.fill = header_fill
                cell.font = white_font
                cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(min_row=6, max_row=max(ws.max_row, 26), min_col=1, max_col=14):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for cell in ["B8", "B9"]:
        ws[cell].number_format = 'R$ #,##0.00'

    ajustar_larguras(ws)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 20
    ws.freeze_panes = "A6"


def salvar_com_indicadores(df_final, lista_renovacao, caminho_saida):
    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        df_final.to_excel(writer, index=False, sheet_name="CONSOLIDADO")
        lista_renovacao.to_excel(writer, index=False, sheet_name="LISTA RENOVACAO")

        wb = writer.book
        ws_consolidado = wb["CONSOLIDADO"]
        ws_renovacao = wb["LISTA RENOVACAO"]

        aplicar_estilo_tabela(ws_consolidado, "TabelaConsolidado")
        aplicar_estilo_tabela(ws_renovacao, "TabelaRenovacao")
        montar_aba_indicadores(writer, df_final)

        ws_renovacao["A1"].comment = None
        if ws_renovacao.max_row > 1:
            for cell in ws_renovacao[1]:
                cell.fill = PatternFill("solid", fgColor="C00000")
                cell.font = Font(color="FFFFFF", bold=True)
            for row in ws_renovacao.iter_rows(min_row=2, max_row=ws_renovacao.max_row):
                row[-1].fill = PatternFill("solid", fgColor="FCE4D6")
                row[-1].font = Font(bold=True, color="9C0006")
        ajustar_larguras(ws_renovacao)


# ============================================================
# PROCESSAMENTO
# ============================================================

arquivos = glob(os.path.join(PASTA, "*.xls")) + glob(os.path.join(PASTA, "*.xlsx"))
arquivos_avp = [a for a in arquivos if nome_arquivo_eh_base_avp(os.path.basename(a))]
arquivos_principais = [a for a in arquivos if a not in arquivos_avp]

lista_df = []
base_avp = carregar_base_avp(arquivos_avp)

print("=" * 70)
print("INICIANDO CONSOLIDACAO")
print("=" * 70)

for arquivo in arquivos_principais:
    nome_arquivo = os.path.basename(arquivo)

    if ARQUIVO_SAIDA.upper() in nome_arquivo.upper():
        continue

    print(f"Lendo arquivo: {nome_arquivo}")

    try:
        df = ler_arquivo(arquivo)

        df = df.dropna(how="all").reset_index(drop=True)

        if df.empty:
            print(f"Arquivo vazio ignorado: {nome_arquivo}")
            continue

        col_nome = localizar_coluna(df, "Nome", 1)
        col_pedido = localizar_coluna(df, "Pedido", 3)

        df[["Nome_Separado", "CPF_CNPJ", "Parceiro"]] = df[col_nome].apply(extrair_nome_cpf_parceiro)
        df[["Pedido_Numero", "Emissor"]] = df[col_pedido].apply(extrair_pedido_emissor)

        df["Arquivo_Origem"] = nome_arquivo

        lista_df.append(df)

        print(f"OK -> {nome_arquivo} | Linhas: {len(df)}")

    except Exception as e:
        print(f"ERRO -> {nome_arquivo}: {e}")

if not lista_df:
    print("=" * 70)
    print("Nenhum arquivo foi processado com sucesso.")
    print("=" * 70)
else:
    df_final = pd.concat(lista_df, ignore_index=True)
    df_final = preparar_campos_indicadores(df_final)
    df_final = aplicar_avp_no_consolidado(df_final, base_avp)
    lista_renovacao = criar_lista_renovacao(df_final)

    saida = os.path.join(PASTA, ARQUIVO_SAIDA)
    try:
        salvar_com_indicadores(df_final, lista_renovacao, saida)
    except PermissionError:
        saida = os.path.join(PASTA, "CONSOLIDADO_FINAL_INDICADORES.xlsx")
        salvar_com_indicadores(df_final, lista_renovacao, saida)

    print("=" * 70)
    print("CONSOLIDADO GERADO COM SUCESSO")
    print(f"Arquivo: {saida}")
    print(f"Total de linhas consolidadas: {len(df_final)}")
    print(f"Clientes sem renovacao 2026: {len(lista_renovacao)}")
    print("Abas adicionadas: INDICADORES COMERCIAIS e LISTA RENOVACAO")
    print("=" * 70)

input("Pressione ENTER para sair...")
